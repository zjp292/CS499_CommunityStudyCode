import requests
import openpyxl

def get_book_genre(title, author):
    # using openlibrary API to get genres
    search_url = 'http://openlibrary.org/search.json'
    params = {
        'title': title,
        'author': author
    }
    response = requests.get(search_url, params=params)

    if response.status_code == 200:
        data = response.json()

        # check if response contains 'docs'
        if data['docs']:
            first_result = data['docs'][0]

            if 'subject' in first_result:
                genres = first_result['subject']
                return genres

            else:
                return "no genres found"

        else:
            return "book was not found"

    else:
        return "failed to find book info"

def get_genre(title, author):
    url = f"http://openlibrary.org/search.json?title={title}&author={author}"
    response = requests.get(url)
    data = response.json()

    if data.get("docs"):
        doc = data["docs"][0]
        return doc.get("subject", [])[:5]

    else:
        return []

def google_genre(title, author):
    url = f'https://www.googleapis.com/books/v1/volumes?q={title}&{author}'
    req = requests.get(url)
    data = req.json()

    if "items" in data and len(data["items"]) > 0:
        item = data["items"][0]
        info = item.get('volumeInfo', {})
        categories = info.get('categories', [])

        if categories:
            return categories
        else:
            return ['no genres']
    else:
        return ['no results']


data = []
path = 'C:/Users/zachp/OneDrive/Desktop/cs499_data_collection/book_data.xlsx'
# open and load the workbook at path
workbook = openpyxl.load_workbook(path)
worksheet = workbook.active

# get titles and authors from spreadsheet
for row in worksheet.iter_rows(values_only=True):
    title, author = row[:2]
    data.append((title, author))



index = 1

for title, author in data:
    genres = get_genre(title, author)
    google_genre = google_genre(title, author)

    # used for open library API
    # if genres:
    #     print(f"title: {title}, author: {author}, genres: {genres}")
    #     worksheet.cell(row=index, column=4).value=', '.join(genres)

    # used for google books api
    if google_genre:
        print(google_genre)
        worksheet.cell(row=index, column=5).value=''.join(google_genre)
    else:
        print(f"Nothing found for title: {title}, author: {author}")

    index = index + 1

workbook.save(path)
