import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

workbook = openpyxl.load_workbook('C:/Users/zachp/OneDrive/Desktop/cs499_data_collection/book_data.xlsx')
worksheet = workbook.active

def main_genre():
    qual_data = []
    # get the main genres
    for row in worksheet.iter_rows(values_only=True):
        qual_data.append(row[4])

    # put data into dataframe
    df = pd.DataFrame(qual_data, columns=['Genre'])

    # format
    count = df['Genre'].value_counts().reset_index()
    count.columns = ['Genre', 'Frequency']

    # sort genres by the frequency that appear
    count = count.sort_values(by='Genre')
    count_sorted = count.sort_values(by='Frequency', ascending=False)

    #pd.set_option('display.max_rows', None)
    #print(count_sorted)
    #print('\n')
    #print(count_sorted.head(15))
    return(count_sorted)

def get_sub_genre():
    sub_genre = []
    # get genre data from data set
    for row in worksheet.iter_rows(values_only=True):
        sub_genre.append((row[4], row[3]))

    df = pd.DataFrame(sub_genre, columns=['Genre', 'Sub-Genre'])
    fiction_analysis = df[df['Genre'] == 'Fiction']

    # split the sub genres by comma
    sub_genre_list = [sub.split(', ') for sub in fiction_analysis['Sub-Genre']]
    all_subgenres = [subgenre.strip() for sublist in sub_genre_list for subgenre in sublist]
    subgenres_df = pd.DataFrame(all_subgenres, columns=['Sub Genre'])

    sub_genre_count = subgenres_df['Sub Genre'].value_counts().reset_index()
    sub_genre_count.columns = ['Sub Genre', 'Frequency']

    most_frequent_subgenres = sub_genre_count.head(13)

    #pd.set_option('display.max_rows', None)
    #print(most_frequent_subgenres)
    return(sub_genre_count)

def descriptive_statistical_analysis():
    genre_freq = main_genre()

    # measures of central tendencies
    mean = genre_freq['Frequency'].mean()
    median = genre_freq['Frequency'].median()
    mode = genre_freq['Frequency'].mode()
    print("mean: ", mean, "\nmedian: ", median, "\nmode: ", mode)

    # measures of variance
    range = genre_freq['Frequency'].max() - genre_freq['Frequency'].min()
    std_dev = genre_freq['Frequency'].std()
    variance = genre_freq['Frequency'].var()
    Q1 = genre_freq['Frequency'].quantile(.25)
    Q3 = genre_freq['Frequency'].quantile(.75)
    IQR = Q3 - Q1
    skew = genre_freq['Frequency'].skew()
    print('\nSkew: ', skew)
    print("\nrange: ", range, "\nstd dev: ", std_dev, "\nvariance: ", variance, '\nIQR: ', IQR)

    # histogram graph
    plt.figure(figsize=(10, 6))
    plt.hist(data=genre_freq, x='Frequency', bins=43)
    plt.title('Distribution of Main Genres')
    plt.ylabel('Frequency')
    plt.xlabel('Count')
    plt.grid(True)
    plt.show()



def main():
    #get_sub_genre()
    #main_genre()
    descriptive_statistical_analysis()


if __name__ == "__main__":
    main()